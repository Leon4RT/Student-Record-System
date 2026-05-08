from flask import Flask, render_template, request, jsonify
import json
import os
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)

# PostgreSQL database URL from environment variable
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://student_records_db_zj1k_user:waMNpFTs2YH0yoGtWuYqfV6c7sHTbEfp@dpg-d7urj6brjlhs7397pr40-a/student_records_db_zj1k')

SUBJECTS = [
    {"code": "CC 101",    "name": "Computer Programming",             "units": 3},
    {"code": "MATH 101",  "name": "Differential Calculus",            "units": 4},
    {"code": "CC 100",    "name": "Introduction to Computing",        "units": 3},
    {"code": "GEd 102",   "name": "Mathematics in the Modern World",  "units": 3},
    {"code": "PATHFit 1", "name": "Movement Competency Training",     "units": 2},
    {"code": "NSTP",      "name": "ROTC/CWTS",                        "units": 3},
    {"code": "GEd 105",   "name": "Readings in Philippine History",   "units": 3},
    {"code": "GEd 101",   "name": "Understanding the Self",           "units": 3},
]

# ── DATABASE SETUP ──────────────────────────────────────────
def get_db():
    """Get a PostgreSQL database connection."""
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def init_db():
    """Create the students table if it doesn't exist."""
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS students (
            student_id  TEXT PRIMARY KEY,
            name        TEXT NOT NULL,
            program     TEXT NOT NULL,
            section     TEXT NOT NULL,
            grades      TEXT NOT NULL DEFAULT '{}',
            gwa         REAL NOT NULL DEFAULT 0.0,
            remarks     TEXT NOT NULL DEFAULT 'N/A'
        )
    ''')
    conn.commit()
    cur.close()
    conn.close()
    print("✅ PostgreSQL Database ready!")

# ── HELPERS ─────────────────────────────────────────────────
def code_key(code):
    """Convert subject code to a safe key."""
    result = ''
    for c in code:
        if c.isalnum() or c == '_':
            result += c
        else:
            result += '_'
    return result

def compute_gwa(grades):
    total_units = 0
    weighted_sum = 0
    for subj in SUBJECTS:
        key = code_key(subj["code"])
        grade = grades.get(key)
        if grade and grade != "":
            try:
                g = float(grade)
                weighted_sum += g * subj["units"]
                total_units += subj["units"]
            except:
                pass
    if total_units == 0:
        return 0.0
    return round(weighted_sum / total_units, 4)

def get_remarks(gwa):
    if gwa == 0:
        return "N/A"
    if gwa <= 1.0:
        return "Excellent"
    elif gwa <= 1.5:
        return "Superior"
    elif gwa <= 2.0:
        return "Very Good"
    elif gwa <= 2.5:
        return "Good"
    elif gwa <= 3.0:
        return "Passing"
    else:
        return "Failed"

def row_to_dict(row):
    """Convert a database row to a dictionary."""
    d = dict(row)
    if isinstance(d['grades'], str):
        d['grades'] = json.loads(d['grades'])
    return d

# ── ROUTES ──────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html', subjects=SUBJECTS)

@app.route('/api/subjects')
def get_subjects():
    return jsonify(SUBJECTS)

@app.route('/api/students', methods=['GET'])
def get_students():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM students ORDER BY name')
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route('/api/students', methods=['POST'])
def add_student():
    data = request.json
    sid  = str(data.get('student_id', '')).strip()
    name = str(data.get('name', '')).strip()
    prog = "BSIT"
    sec  = str(data.get('section', '')).strip()

    if not sid:
        return jsonify({"error": "Student ID is required"}), 400
    if not name:
        return jsonify({"error": "Name is required"}), 400
    if not sec:
        return jsonify({"error": "Section is required"}), 400

    grades  = data.get('grades', {})
    gwa     = compute_gwa(grades)
    remarks = get_remarks(gwa)

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO students (student_id, name, program, section, grades, gwa, remarks) VALUES (%s, %s, %s, %s, %s, %s, %s)',
            (sid, name, prog, sec, json.dumps(grades), gwa, remarks)
        )
        conn.commit()
        cur.close()
        conn.close()
    except psycopg2.IntegrityError:
        return jsonify({"error": "Student ID already exists"}), 400

    return jsonify({
        "student_id": sid, "name": name, "program": prog,
        "section": sec, "grades": grades, "gwa": gwa, "remarks": remarks
    }), 201

@app.route('/api/students/<sid>', methods=['GET'])
def get_student(sid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM students WHERE student_id = %s', (sid,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return jsonify({"error": "Student not found"}), 404
    return jsonify(row_to_dict(row))

@app.route('/api/students/<sid>', methods=['PUT'])
def update_student(sid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM students WHERE student_id = %s', (sid,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"error": "Student not found"}), 404

    data    = request.json
    student = row_to_dict(row)
    name    = data.get('name', student['name'])
    prog    = "BSIT"
    sec     = data.get('section', student['section'])
    grades  = data.get('grades', student['grades'])
    gwa     = compute_gwa(grades)
    remarks = get_remarks(gwa)

    cur.execute(
        'UPDATE students SET name=%s, program=%s, section=%s, grades=%s, gwa=%s, remarks=%s WHERE student_id=%s',
        (name, prog, sec, json.dumps(grades), gwa, remarks, sid)
    )
    conn.commit()
    cur.close()
    conn.close()

    return jsonify({
        "student_id": sid, "name": name, "program": prog,
        "section": sec, "grades": grades, "gwa": gwa, "remarks": remarks
    })

@app.route('/api/students/<sid>', methods=['DELETE'])
def delete_student(sid):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM students WHERE student_id = %s', (sid,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"error": "Student not found"}), 404
    cur.execute('DELETE FROM students WHERE student_id = %s', (sid,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"success": True})


@app.route('/api/export/excel')
def export_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file

    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM students ORDER BY program, section, name')
    rows = cur.fetchall()
    cur.close()
    conn.close()
    students_list = [row_to_dict(r) for r in rows]

    wb = Workbook()
    wb.remove(wb.active)

    # Colors
    RED       = 'C0152A'
    RED_LIGHT = 'FFE8E8'
    WHITE     = 'FFFFFF'
    GRAY      = 'F5F5F5'
    DARK      = '2C2C2C'

    def make_border():
        thin = Side(style='thin', color='DDDDDD')
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_cell(ws, row, col, value, bg=RED, fg=WHITE, bold=True, size=11, center=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fg, size=size, name='Arial')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
        c.border = make_border()
        return c

    def data_cell(ws, row, col, value, bold=False, center=False, bg=WHITE):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=DARK, size=10, name='Arial')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
        c.border = make_border()
        return c

    # ── SHEET 1: ALL STUDENTS ──────────────────────────────
    ws_all = wb.create_sheet('All Students')
    ws_all.sheet_view.showGridLines = False

    # Title
    ws_all.merge_cells('A1:O1')
    t = ws_all['A1']
    t.value = 'BSIT 1ST YEAR — COMPLETE STUDENT RECORDS'
    t.font = Font(bold=True, color=WHITE, size=14, name='Arial')
    t.fill = PatternFill('solid', start_color=RED)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws_all.row_dimensions[1].height = 36

    # Sub-title
    ws_all.merge_cells('A2:O2')
    s = ws_all['A2']
    s.value = f'Total Students: {len(students_list)}'
    s.font = Font(bold=True, color=RED, size=11, name='Arial')
    s.fill = PatternFill('solid', start_color=RED_LIGHT)
    s.alignment = Alignment(horizontal='center', vertical='center')
    ws_all.row_dimensions[2].height = 22

    # Headers row
    headers = ['#', 'Student ID', 'Name', 'Program', 'Section',
               'CC 101', 'MATH 101', 'CC 100', 'GEd 102',
               'PATHFit 1', 'NSTP', 'GEd 105', 'GEd 101',
               'GWA', 'Remarks']
    ws_all.row_dimensions[3].height = 38
    for ci, h in enumerate(headers, 1):
        header_cell(ws_all, 3, ci, h)

    subj_keys = ['CC_101','MATH_101','CC_100','GEd_102','PATHFit_1','NSTP','GEd_105','GEd_101']

    for i, s in enumerate(students_list, 1):
        r = i + 3
        bg = WHITE if i % 2 == 1 else GRAY
        ws_all.row_dimensions[r].height = 20
        data_cell(ws_all, r, 1, i, center=True, bg=bg)
        data_cell(ws_all, r, 2, s['student_id'], center=True, bg=bg)
        data_cell(ws_all, r, 3, s['name'], bg=bg)
        data_cell(ws_all, r, 4, s['program'], center=True, bg=bg)
        data_cell(ws_all, r, 5, s['section'], center=True, bg=bg)
        for ki, key in enumerate(subj_keys):
            val = s['grades'].get(key, '—')
            data_cell(ws_all, r, 6+ki, val, center=True, bg=bg)
        gwa_cell = data_cell(ws_all, r, 14, round(s['gwa'], 4) if s['gwa'] else '—', bold=True, center=True, bg=bg)
        rem_cell = data_cell(ws_all, r, 15, s['remarks'], center=True, bg=bg)

        # Color remarks
        rem_colors = {'PASSED':'155724','FAILED':'721c24','N/A':'666666'}
        rem_bg     = {'PASSED':'d4edda','FAILED':'f8d7da','N/A':'F5F5F5'}
        rk = s['remarks']
        if rk in rem_colors:
            rem_cell.font = Font(bold=True, color=rem_colors[rk], size=10, name='Arial')
            rem_cell.fill = PatternFill('solid', start_color=rem_bg[rk])

    # Column widths
    col_widths = [4, 14, 26, 10, 10, 9, 10, 9, 10, 10, 8, 10, 10, 8, 12]
    for ci, w in enumerate(col_widths, 1):
        ws_all.column_dimensions[get_column_letter(ci)].width = w

    # ── SHEETS BY PROGRAM + SECTION ───────────────────────
    from itertools import groupby

    # Group by program first, then section
    by_prog = {}
    for s in students_list:
        key = s['program'] or 'Unknown'
        by_prog.setdefault(key, []).append(s)

    for prog, prog_students in sorted(by_prog.items()):
        by_sec = {}
        for s in prog_students:
            key = s['section'] or 'Unknown'
            by_sec.setdefault(key, []).append(s)

        for sec, sec_students in sorted(by_sec.items()):
            sheet_name = f'Section {sec}'[:31]
            ws = wb.create_sheet(sheet_name)
            ws.sheet_view.showGridLines = False

            # Title
            ws.merge_cells('A1:O1')
            t = ws['A1']
            t.value = f'BSIT 1ST YEAR — SECTION {sec}'
            t.font = Font(bold=True, color=WHITE, size=13, name='Arial')
            t.fill = PatternFill('solid', start_color=RED)
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 34

            ws.merge_cells('A2:O2')
            s2 = ws['A2']
            s2.value = f'Program: BSIT  |  Year Level: 1st Year  |  Section: {sec}  |  Total Students: {len(sec_students)}'
            s2.font = Font(bold=True, color=RED, size=10, name='Arial')
            s2.fill = PatternFill('solid', start_color=RED_LIGHT)
            s2.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 20

            ws.row_dimensions[3].height = 38
            for ci, h in enumerate(headers, 1):
                header_cell(ws, 3, ci, h)

            for i, st in enumerate(sec_students, 1):
                r = i + 3
                bg = WHITE if i % 2 == 1 else GRAY
                ws.row_dimensions[r].height = 20
                data_cell(ws, r, 1, i, center=True, bg=bg)
                data_cell(ws, r, 2, st['student_id'], center=True, bg=bg)
                data_cell(ws, r, 3, st['name'], bg=bg)
                data_cell(ws, r, 4, st['program'], center=True, bg=bg)
                data_cell(ws, r, 5, st['section'], center=True, bg=bg)
                for ki, key in enumerate(subj_keys):
                    val = st['grades'].get(key, '—')
                    data_cell(ws, r, 6+ki, val, center=True, bg=bg)
                data_cell(ws, r, 14, round(st['gwa'], 4) if st['gwa'] else '—', bold=True, center=True, bg=bg)
                rc = data_cell(ws, r, 15, st['remarks'], center=True, bg=bg)
                rk = st['remarks']
                if rk in rem_colors:
                    rc.font = Font(bold=True, color=rem_colors[rk], size=10, name='Arial')
                    rc.fill = PatternFill('solid', start_color=rem_bg[rk])

            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            # Section GWA summary row
            total_rows = len(sec_students) + 3
            ws.merge_cells(f'A{total_rows+1}:M{total_rows+1}')
            avg_cell = ws.cell(row=total_rows+1, column=14)
            avg_cell.value = f'=AVERAGE(N4:N{total_rows})'
            avg_cell.font = Font(bold=True, color=WHITE, size=11, name='Arial')
            avg_cell.fill = PatternFill('solid', start_color=RED)
            avg_cell.alignment = Alignment(horizontal='center', vertical='center')
            avg_cell.border = make_border()

            lbl = ws.cell(row=total_rows+1, column=1)
            ws.merge_cells(f'A{total_rows+1}:M{total_rows+1}')
            lbl.value = 'Section Average GWA:'
            lbl.font = Font(bold=True, color=WHITE, size=11, name='Arial')
            lbl.fill = PatternFill('solid', start_color=RED)
            lbl.alignment = Alignment(horizontal='right', vertical='center')
            lbl.border = make_border()
            ws.row_dimensions[total_rows+1].height = 24

    # Save to memory
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Student_Records.xlsx'
    )

# ── START ────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    print("🎓 Student Record System running at http://127.0.0.1:5000")
    import os
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)